from openpyxl import load_workbook


def fetchUserList(localPath):
    # path of excel file
    wb = load_workbook(localPath)
    ws = wb.active

    # to find max value of row for a col
    rowcount = 0
    for cell in ws["A"]:
        if not cell.value is None:
            rowcount += 1

    # to find max value of col for a row
    colCount = 0
    for cell in ws[1]:
        if not cell.value is None:
            colCount += 1

    for row in ws.iter_rows(min_row=2, max_row=rowcount, min_col=1, max_col=colCount):
        inList = []
        for cell in row:
            inList.append(cell.value)

    return inList
